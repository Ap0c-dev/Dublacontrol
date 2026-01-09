#!/usr/bin/env python3
"""
Script para importar alunos de arquivo Excel diretamente no banco de dados
"""

import sys
import os

# Adicionar o diretório raiz ao path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import create_app, db
from app.models.aluno import Aluno
from app.models.professor import Professor
from app.models.matricula import Matricula
from app.models.horario_professor import HorarioProfessor
from app.routes import normalizar_texto
from openpyxl import load_workbook
from datetime import datetime, date
from calendar import monthrange

def importar_alunos(arquivo_excel):
    """Importa alunos de arquivo Excel"""
    
    app = create_app()
    
    with app.app_context():
        print("📂 Carregando arquivo Excel...")
        
        try:
            wb = load_workbook(arquivo_excel, data_only=True)
            ws = wb.active
        except Exception as e:
            print(f"❌ Erro ao ler arquivo Excel: {e}")
            return
        
        # Ler cabeçalhos
        headers = []
        for cell in ws[1]:
            headers.append(cell.value.lower().strip() if cell.value else '')
        
        print(f"📋 Cabeçalhos encontrados: {len(headers)} colunas")
        
        # Mapear nomes de colunas
        def encontrar_coluna(nomes_possiveis):
            for nome in nomes_possiveis:
                for idx, header in enumerate(headers):
                    if nome.lower() in header.lower():
                        return idx
            return None
        
        # Índices das colunas
        idx_nome = encontrar_coluna(['nome', 'name'])
        idx_telefone = encontrar_coluna(['telefone', 'phone', 'tel'])
        idx_cidade = encontrar_coluna(['cidade', 'city'])
        idx_estado = encontrar_coluna(['estado', 'state', 'uf'])
        idx_forma_pagamento = encontrar_coluna(['forma_pagamento', 'forma pagamento', 'pagamento', 'payment'])
        idx_data_vencimento = encontrar_coluna(['data_vencimento', 'data vencimento', 'vencimento', 'due_date'])
        idx_nome_responsavel = encontrar_coluna(['nome_responsavel', 'nome responsavel', 'responsavel', 'responsible'])
        idx_telefone_responsavel = encontrar_coluna(['telefone_responsavel', 'telefone responsavel', 'tel responsavel'])
        idx_data_nascimento = encontrar_coluna(['data_nascimento', 'data nascimento', 'nascimento', 'birth', 'birthday'])
        
        # Modalidades
        idx_dublagem_online = encontrar_coluna(['dublagem_online', 'dublagem online'])
        idx_dublagem_presencial = encontrar_coluna(['dublagem_presencial', 'dublagem presencial'])
        idx_teatro_online = encontrar_coluna(['teatro_online', 'teatro online'])
        idx_teatro_presencial = encontrar_coluna(['teatro_presencial', 'teatro presencial'])
        idx_locucao = encontrar_coluna(['locucao', 'locução'])
        idx_teatro_tv_cinema = encontrar_coluna(['teatro_tv_cinema', 'teatro tv cinema', 'tv cinema'])
        idx_musical = encontrar_coluna(['musical'])
        
        # Status
        idx_aprovado = encontrar_coluna(['aprovado', 'approved'])
        idx_ativo = encontrar_coluna(['ativo', 'active'])
        idx_experimental = encontrar_coluna(['experimental'])
        
        # Professor/Matrícula
        idx_professor_nome = encontrar_coluna(['professor_nome', 'professor', 'nome_professor'])
        idx_professor_modalidade = encontrar_coluna(['professor_modalidade', 'modalidade', 'tipo_curso', 'curso'])
        idx_valor_mensalidade = encontrar_coluna(['valor_mensalidade', 'mensalidade', 'valor'])
        idx_data_inicio = encontrar_coluna(['data_inicio', 'data inicio', 'inicio'])
        idx_horario_dia_semana = encontrar_coluna(['horario_dia_semana', 'dia_semana', 'dia semana', 'dia'])
        idx_horario_aula = encontrar_coluna(['horario_aula', 'horario', 'horário'])
        
        # Validar colunas obrigatórias
        erros = []
        if idx_nome is None:
            erros.append('Coluna "nome" não encontrada')
        if idx_telefone is None:
            erros.append('Coluna "telefone" não encontrada')
        if idx_cidade is None:
            erros.append('Coluna "cidade" não encontrada')
        if idx_estado is None:
            erros.append('Coluna "estado" não encontrada')
        if idx_forma_pagamento is None:
            erros.append('Coluna "forma_pagamento" não encontrada')
        if idx_data_vencimento is None:
            erros.append('Coluna "data_vencimento" não encontrada')
        
        if erros:
            print("❌ Erros encontrados:")
            for erro in erros:
                print(f"  - {erro}")
            return
        
        # Processar linhas
        alunos_criados = 0
        alunos_erro = []
        alunos_duplicados = []
        matriculas_criadas = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if not row[idx_nome].value:
                continue
            
            try:
                # Ler valores básicos
                nome = str(row[idx_nome].value).strip() if row[idx_nome].value else ''
                telefone = str(row[idx_telefone].value).strip() if row[idx_telefone].value else ''
                cidade = str(row[idx_cidade].value).strip() if row[idx_cidade].value else ''
                estado = str(row[idx_estado].value).strip().upper() if row[idx_estado].value else ''
                forma_pagamento = str(row[idx_forma_pagamento].value).strip() if row[idx_forma_pagamento].value else ''
                
                if not nome or not telefone or not cidade or not estado or not forma_pagamento:
                    alunos_erro.append({
                        'linha': row_idx,
                        'nome': nome or '(vazio)',
                        'erro': 'Campos obrigatórios faltando'
                    })
                    continue
                
                # Verificar duplicado
                aluno_existente = Aluno.query.filter_by(telefone=telefone, ativo=True).first()
                if aluno_existente:
                    alunos_duplicados.append({
                        'linha': row_idx,
                        'nome': nome,
                        'telefone': telefone
                    })
                    continue
                
                # Normalizar
                nome = normalizar_texto(nome)
                cidade = normalizar_texto(cidade)
                
                # Processar data_vencimento
                data_vencimento = None
                if row[idx_data_vencimento].value:
                    valor_venc = row[idx_data_vencimento].value
                    if isinstance(valor_venc, datetime):
                        data_vencimento = valor_venc.date()
                    elif isinstance(valor_venc, date):
                        data_vencimento = valor_venc
                    elif isinstance(valor_venc, (int, float)):
                        dia = int(valor_venc)
                        if 1 <= dia <= 31:
                            hoje = date.today()
                            try:
                                data_vencimento = date(hoje.year, hoje.month, dia)
                            except ValueError:
                                ultimo_dia = monthrange(hoje.year, hoje.month)[1]
                                data_vencimento = date(hoje.year, hoje.month, min(dia, ultimo_dia))
                    else:
                        try:
                            data_vencimento = datetime.strptime(str(valor_venc), '%d/%m/%Y').date()
                        except:
                            try:
                                data_vencimento = datetime.strptime(str(valor_venc), '%Y-%m-%d').date()
                            except:
                                pass
                
                if not data_vencimento:
                    hoje = date.today()
                    data_vencimento = date(hoje.year, hoje.month, 10)
                
                # Processar data_nascimento
                data_nascimento = None
                if idx_data_nascimento is not None and row[idx_data_nascimento].value:
                    valor_nasc = row[idx_data_nascimento].value
                    if isinstance(valor_nasc, datetime):
                        data_nascimento = valor_nasc.date()
                    elif isinstance(valor_nasc, date):
                        data_nascimento = valor_nasc
                    else:
                        try:
                            data_nascimento = datetime.strptime(str(valor_nasc), '%d/%m/%Y').date()
                        except:
                            try:
                                data_nascimento = datetime.strptime(str(valor_nasc), '%Y-%m-%d').date()
                            except:
                                pass
                
                # Campos opcionais
                nome_responsavel = None
                if idx_nome_responsavel is not None and row[idx_nome_responsavel].value:
                    nome_responsavel = normalizar_texto(str(row[idx_nome_responsavel].value).strip())
                    if not nome_responsavel:
                        nome_responsavel = None
                
                telefone_responsavel = None
                if idx_telefone_responsavel is not None and row[idx_telefone_responsavel].value:
                    telefone_responsavel = str(row[idx_telefone_responsavel].value).strip()
                    if not telefone_responsavel:
                        telefone_responsavel = None
                
                # Modalidades
                def processar_boolean(idx):
                    if idx is not None and row[idx].value:
                        valor = str(row[idx].value).lower().strip()
                        return valor in ['sim', 's', 'yes', 'y', '1', 'true', 'x', '✓']
                    return False
                
                dublagem_online = processar_boolean(idx_dublagem_online)
                dublagem_presencial = processar_boolean(idx_dublagem_presencial)
                teatro_online = processar_boolean(idx_teatro_online)
                teatro_presencial = processar_boolean(idx_teatro_presencial)
                locucao = processar_boolean(idx_locucao)
                teatro_tv_cinema = processar_boolean(idx_teatro_tv_cinema)
                musical = processar_boolean(idx_musical)
                
                # Status
                aprovado = True
                if idx_aprovado is not None and row[idx_aprovado].value:
                    valor = str(row[idx_aprovado].value).lower().strip()
                    aprovado = valor in ['sim', 's', 'yes', 'y', '1', 'true', 'aprovado']
                
                ativo = True
                if idx_ativo is not None and row[idx_ativo].value:
                    valor = str(row[idx_ativo].value).lower().strip()
                    ativo = valor in ['sim', 's', 'yes', 'y', '1', 'true', 'ativo']
                
                experimental = False
                if idx_experimental is not None and row[idx_experimental].value:
                    valor = str(row[idx_experimental].value).lower().strip()
                    experimental = valor in ['sim', 's', 'yes', 'y', '1', 'true', 'experimental']
                
                # Criar aluno
                aluno = Aluno(
                    nome=nome,
                    telefone=telefone,
                    nome_responsavel=nome_responsavel,
                    telefone_responsavel=telefone_responsavel,
                    cidade=cidade,
                    estado=estado,
                    forma_pagamento=forma_pagamento,
                    data_vencimento=data_vencimento,
                    data_nascimento=data_nascimento,
                    dublagem_online=dublagem_online,
                    dublagem_presencial=dublagem_presencial,
                    teatro_online=teatro_online,
                    teatro_presencial=teatro_presencial,
                    locucao=locucao,
                    teatro_tv_cinema=teatro_tv_cinema,
                    musical=musical,
                    aprovado=aprovado,
                    ativo=ativo,
                    experimental=experimental
                )
                
                db.session.add(aluno)
                db.session.flush()
                
                if data_nascimento:
                    aluno.idade = aluno.calcular_idade()
                
                # Criar matrícula se houver professor
                if idx_professor_nome is not None and row[idx_professor_nome].value:
                    professor_nome = str(row[idx_professor_nome].value).strip()
                    if professor_nome:
                        professor_nome_normalizado = normalizar_texto(professor_nome)
                        professor = Professor.query.filter(
                            db.func.lower(Professor.nome) == professor_nome_normalizado.lower(),
                            Professor.ativo == True
                        ).first()
                        
                        if professor:
                            # Obter modalidade
                            modalidade = None
                            if idx_professor_modalidade is not None and row[idx_professor_modalidade].value:
                                modalidade_str = str(row[idx_professor_modalidade].value).strip().lower()
                                modalidades_map = {
                                    'dublagem online': 'dublagem_online',
                                    'dublagem_online': 'dublagem_online',
                                    'dublagem presencial': 'dublagem_presencial',
                                    'dublagem_presencial': 'dublagem_presencial',
                                    'teatro online': 'teatro_online',
                                    'teatro_online': 'teatro_online',
                                    'teatro presencial': 'teatro_presencial',
                                    'teatro_presencial': 'teatro_presencial',
                                    'locucao': 'locucao',
                                    'locução': 'locucao',
                                    'teatro tv cinema': 'teatro_tv_cinema',
                                    'teatro_tv_cinema': 'teatro_tv_cinema',
                                    'tv cinema': 'teatro_tv_cinema',
                                    'musical': 'musical'
                                }
                                modalidade = modalidades_map.get(modalidade_str)
                            
                            if not modalidade:
                                if dublagem_online:
                                    modalidade = 'dublagem_online'
                                elif dublagem_presencial:
                                    modalidade = 'dublagem_presencial'
                                elif teatro_online:
                                    modalidade = 'teatro_online'
                                elif teatro_presencial:
                                    modalidade = 'teatro_presencial'
                                elif locucao:
                                    modalidade = 'locucao'
                                elif teatro_tv_cinema:
                                    modalidade = 'teatro_tv_cinema'
                                elif musical:
                                    modalidade = 'musical'
                            
                            if modalidade:
                                # Valor mensalidade
                                valor_mensalidade = None
                                if idx_valor_mensalidade is not None and row[idx_valor_mensalidade].value:
                                    try:
                                        valor_str = str(row[idx_valor_mensalidade].value).replace(',', '.').strip()
                                        valor_mensalidade = float(valor_str)
                                    except (ValueError, AttributeError):
                                        pass
                                
                                # Data início
                                data_inicio_matricula = None
                                if idx_data_inicio is not None and row[idx_data_inicio].value:
                                    valor_inicio = row[idx_data_inicio].value
                                    if isinstance(valor_inicio, datetime):
                                        data_inicio_matricula = valor_inicio.date()
                                    elif isinstance(valor_inicio, date):
                                        data_inicio_matricula = valor_inicio
                                    else:
                                        try:
                                            data_inicio_matricula = datetime.strptime(str(valor_inicio), '%d/%m/%Y').date()
                                        except:
                                            try:
                                                data_inicio_matricula = datetime.strptime(str(valor_inicio), '%Y-%m-%d').date()
                                            except:
                                                pass
                                
                                # Horário
                                dia_semana = None
                                horario_aula = None
                                
                                if idx_horario_dia_semana is not None and row[idx_horario_dia_semana].value:
                                    dia_semana = str(row[idx_horario_dia_semana].value).strip()
                                
                                if idx_horario_aula is not None and row[idx_horario_aula].value:
                                    horario_aula = str(row[idx_horario_aula].value).strip()
                                
                                # Buscar horário automaticamente se não informado
                                if not dia_semana or not horario_aula:
                                    horario_professor = HorarioProfessor.query.filter_by(
                                        professor_id=professor.id,
                                        modalidade=modalidade
                                    ).first()
                                    
                                    if horario_professor:
                                        if not dia_semana:
                                            dia_semana = horario_professor.dia_semana
                                        if not horario_aula:
                                            horario_aula = horario_professor.horario_aula
                                
                                # Criar matrícula
                                matricula = Matricula(
                                    aluno_id=aluno.id,
                                    professor_id=professor.id,
                                    tipo_curso=modalidade,
                                    valor_mensalidade=valor_mensalidade,
                                    data_inicio=data_inicio_matricula,
                                    dia_semana=dia_semana,
                                    horario_aula=horario_aula
                                )
                                db.session.add(matricula)
                                matriculas_criadas += 1
                                print(f"  ✅ Matrícula criada: {professor.nome} - {modalidade}")
                
                alunos_criados += 1
                print(f"✅ Aluno criado (linha {row_idx}): {nome}")
                
            except Exception as e:
                import traceback
                print(f"❌ Erro na linha {row_idx}: {str(e)}")
                alunos_erro.append({
                    'linha': row_idx,
                    'nome': str(row[idx_nome].value) if row[idx_nome].value else '(vazio)',
                    'erro': str(e)
                })
                continue
        
        # Commit
        try:
            db.session.commit()
            print(f"\n✅ Importação concluída!")
            print(f"  📊 Alunos criados: {alunos_criados}")
            print(f"  📚 Matrículas criadas: {matriculas_criadas}")
            if alunos_duplicados:
                print(f"  ⚠️  Alunos duplicados (ignorados): {len(alunos_duplicados)}")
            if alunos_erro:
                print(f"  ❌ Erros: {len(alunos_erro)}")
                for erro in alunos_erro:
                    print(f"    - Linha {erro['linha']}: {erro['nome']} - {erro['erro']}")
        except Exception as e:
            db.session.rollback()
            print(f"❌ Erro ao salvar no banco: {str(e)}")
            import traceback
            print(traceback.format_exc())

if __name__ == '__main__':
    arquivo = 'exemplo_importacao_alunos.xlsx'
    
    if len(sys.argv) > 1:
        arquivo = sys.argv[1]
    
    if not os.path.exists(arquivo):
        print(f"❌ Arquivo não encontrado: {arquivo}")
        sys.exit(1)
    
    print(f"📂 Importando alunos de: {arquivo}")
    importar_alunos(arquivo)

