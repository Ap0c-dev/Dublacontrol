#!/usr/bin/env python3
"""
Script para criar arquivo Excel de exemplo para importação de alunos
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, datetime

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Alunos"

# Definir cabeçalhos
headers = [
    "nome",
    "telefone",
    "cidade",
    "estado",
    "forma_pagamento",
    "data_vencimento",
    "nome_responsavel",
    "telefone_responsavel",
    "data_nascimento",
    "dublagem_online",
    "dublagem_presencial",
    "teatro_online",
    "teatro_presencial",
    "locucao",
    "teatro_tv_cinema",
    "musical",
    "aprovado",
    "ativo",
    "experimental",
    "professor_nome",
    "professor_modalidade",
    "valor_mensalidade",
    "data_inicio",
    "horario_dia_semana",
    "horario_aula"
]

# Adicionar cabeçalhos
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Dados de exemplo
exemplos = [
    {
        "nome": "João Silva",
        "telefone": "(11) 98765-4321",
        "cidade": "São Paulo",
        "estado": "SP",
        "forma_pagamento": "PIX",
        "data_vencimento": 10,  # Dia 10 do mês
        "nome_responsavel": "Maria Silva",
        "telefone_responsavel": "(11) 91234-5678",
        "data_nascimento": date(2010, 5, 15),
        "dublagem_online": "Sim",
        "dublagem_presencial": "Não",
        "teatro_online": "Não",
        "teatro_presencial": "Sim",
        "locucao": "Não",
        "teatro_tv_cinema": "Não",
        "musical": "Não",
        "aprovado": "Sim",
        "ativo": "Sim",
        "experimental": "Não",
        "professor_nome": "Prof. Carlos",
        "professor_modalidade": "dublagem_online",
        "valor_mensalidade": 150.00,
        "data_inicio": date(2024, 1, 15),
        "horario_dia_semana": "Segunda-feira",
        "horario_aula": "17:00 às 19:00"
    },
    {
        "nome": "Ana Costa",
        "telefone": "(21) 99876-5432",
        "cidade": "Rio de Janeiro",
        "estado": "RJ",
        "forma_pagamento": "Boleto",
        "data_vencimento": 15,  # Dia 15 do mês
        "nome_responsavel": None,
        "telefone_responsavel": None,
        "data_nascimento": date(2008, 3, 20),
        "dublagem_online": "Sim",
        "dublagem_presencial": "Sim",
        "teatro_online": "Não",
        "teatro_presencial": "Não",
        "locucao": "Sim",
        "teatro_tv_cinema": "Não",
        "musical": "Não",
        "aprovado": "Sim",
        "ativo": "Sim",
        "experimental": "Não",
        "professor_nome": "Prof. Ana Paula",
        "professor_modalidade": "locucao",
        "valor_mensalidade": 200.00,
        "data_inicio": date(2024, 2, 1),
        "horario_dia_semana": "Quarta-feira",
        "horario_aula": "19:00 às 21:00"
    },
    {
        "nome": "Pedro Santos",
        "telefone": "(31) 97654-3210",
        "cidade": "Belo Horizonte",
        "estado": "MG",
        "forma_pagamento": "Cartão de Crédito",
        "data_vencimento": date(2024, 12, 5),  # Data completa
        "nome_responsavel": None,
        "telefone_responsavel": None,
        "data_nascimento": date(2012, 8, 10),
        "dublagem_online": "Não",
        "dublagem_presencial": "Sim",
        "teatro_online": "Não",
        "teatro_presencial": "Sim",
        "locucao": "Não",
        "teatro_tv_cinema": "Sim",
        "musical": "Sim",
        "aprovado": "Sim",
        "ativo": "Sim",
        "experimental": "Não"
    },
    {
        "nome": "Maria Oliveira",
        "telefone": "(41) 96543-2109",
        "cidade": "Curitiba",
        "estado": "PR",
        "forma_pagamento": "PIX",
        "data_vencimento": 20,  # Dia 20 do mês
        "nome_responsavel": "José Oliveira",
        "telefone_responsavel": "(41) 91234-5678",
        "data_nascimento": date(2011, 11, 25),
        "dublagem_online": "Sim",
        "dublagem_presencial": "Não",
        "teatro_online": "Sim",
        "teatro_presencial": "Não",
        "locucao": "Não",
        "teatro_tv_cinema": "Não",
        "musical": "Não",
        "aprovado": "Sim",
        "ativo": "Sim",
        "experimental": "Sim"  # Aluno experimental
    }
]

# Adicionar dados de exemplo
for row_idx, exemplo in enumerate(exemplos, start=2):
    for col_idx, header in enumerate(headers, start=1):
        valor = exemplo.get(header)
        
        # Tratar valores None
        if valor is None:
            valor = ""
        # Tratar datas
        elif isinstance(valor, date):
            valor = valor.strftime("%d/%m/%Y")
        # Tratar números (dia do mês)
        elif isinstance(valor, int) and header == "data_vencimento":
            # Manter como número para mostrar que pode ser apenas o dia
            pass
        
        ws.cell(row=row_idx, column=col_idx, value=valor)

# Ajustar largura das colunas
column_widths = {
    'A': 20,  # nome
    'B': 18,  # telefone
    'C': 15,  # cidade
    'D': 8,   # estado
    'E': 18,  # forma_pagamento
    'F': 15,  # data_vencimento
    'G': 20,  # nome_responsavel
    'H': 18,  # telefone_responsavel
    'I': 15,  # data_nascimento
    'J': 15,  # dublagem_online
    'K': 18,  # dublagem_presencial
    'L': 15,  # teatro_online
    'M': 18,  # teatro_presencial
    'N': 12,  # locucao
    'O': 18,  # teatro_tv_cinema
    'P': 12,  # musical
    'Q': 12,  # aprovado
    'R': 10,  # ativo
    'S': 15,  # experimental
    'T': 18,  # professor_nome
    'U': 20,  # professor_modalidade
    'V': 18,  # valor_mensalidade
    'W': 15,  # data_inicio
    'X': 18,  # horario_dia_semana
    'Y': 18   # horario_aula
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Adicionar linha de instruções (linha após os dados)
linha_instrucoes = len(exemplos) + 3
ws.cell(row=linha_instrucoes, column=1, value="INSTRUÇÕES:")
ws.cell(row=linha_instrucoes, column=1).font = Font(bold=True)

instrucoes = [
    "1. Colunas obrigatórias: nome, telefone, cidade, estado, forma_pagamento, data_vencimento",
    "2. data_vencimento pode ser: número (dia do mês) ou data completa (DD/MM/YYYY)",
    "3. Para modalidades e status, use: Sim/S/Não/N ou 1/0 ou X/✓",
    "4. Alunos com mesmo telefone serão ignorados (duplicados)",
    "5. Campos opcionais podem ficar em branco",
    "6. A primeira linha deve conter os cabeçalhos (não altere os nomes das colunas)",
    "",
    "COLUNAS DE PROFESSOR (opcionais):",
    "- professor_nome: Nome do professor (deve existir no sistema)",
    "- professor_modalidade: Modalidade do curso (dublagem_online, dublagem_presencial, etc)",
    "- valor_mensalidade: Valor da mensalidade (opcional)",
    "- data_inicio: Data de início da matrícula (DD/MM/YYYY, opcional)",
    "- horario_dia_semana: Dia da semana (ex: Segunda-feira, Terça-feira, opcional)",
    "- horario_aula: Horário da aula (ex: 17:00 às 19:00, opcional)",
    "",
    "NOTA: Se informar professor_nome, o sistema criará automaticamente a matrícula do aluno com o professor.",
    "Se não informar horario_dia_semana e horario_aula, o sistema tentará buscar automaticamente",
    "o primeiro horário disponível do professor para aquela modalidade."
]

for idx, instrucao in enumerate(instrucoes, start=linha_instrucoes + 1):
    ws.cell(row=idx, column=1, value=instrucao)

# Salvar arquivo
nome_arquivo = "exemplo_importacao_alunos.xlsx"
wb.save(nome_arquivo)
print(f"✅ Arquivo Excel criado: {nome_arquivo}")
print(f"📊 Total de colunas: {len(headers)}")
print(f"📝 Total de exemplos: {len(exemplos)}")

